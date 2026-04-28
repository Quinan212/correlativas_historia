import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Intentamos encontrar las hojas por coincidencia parcial si el nombre exacto falla por tildes
    target_names = ['Historia', 'Geografía', 'Ciencias Políticas']
    actual_sheets = wb.sheetnames
    
    print("\n--- ESCANEO DE FECHAS EN PESTAÑAS REALES ---")
    
    for target in target_names:
        found_sheet = None
        for sname in actual_sheets:
            if target.lower() in sname.lower():
                found_sheet = sname
                break
        
        if not found_sheet:
            print(f"No se encontró la hoja para: {target}")
            continue
            
        print(f"\nAUDITANDO HOJA: {found_sheet}")
        ws = wb[found_sheet]
        
        for r in range(1, 150):
            row_vals = []
            has_date = False
            for c in range(1, 15):
                val = ws.cell(row=r, column=c).value
                if val:
                    val_str = str(val)
                    row_vals.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val_str}")
                    if "2026" in val_str or (isinstance(val, (int, float)) and 45000 < val < 46500): # Fechas Excel 2026
                        has_date = True
            
            if has_date:
                print(" | ".join(row_vals))

except Exception as e:
    print(f"Error: {e}")
