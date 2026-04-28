import pandas as pd
from openpyxl import load_workbook

file_path = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta\Mesas Extraordinarias de Mayo Ciclo 2026 (1).xlsx'

try:
    # Cargar el Excel para ver las hojas
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    print(f"Hojas encontradas: {sheets}")

    for sheet in sheets:
        print(f"\n--- Analizando Hoja: {sheet} ---")
        ws = wb[sheet]
        
        # Ver las primeras 20 filas y 10 columnas
        data = []
        for row in range(1, 21):
            row_data = []
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        
        df = pd.DataFrame(data)
        print(df.to_string())

except Exception as e:
    print(f"Error: {e}")
