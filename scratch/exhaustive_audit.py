import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta\Mesas Extraordinarias de Mayo Ciclo 2026 (1).xlsx'

def analyze_sheet(ws):
    data = []
    # Leemos absolutamente todo lo que tenga datos según el max_row/max_col de openpyxl
    for row in ws.iter_rows(values_only=True):
        # Solo guardamos la fila si tiene al menos una celda con contenido
        if any(cell is not None for cell in row):
            data.append([str(c) if c is not None else "" for c in row])
    return data

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    full_audit = {}
    
    for sheet_name in wb.sheetnames:
        print(f"Leyendo hoja: {sheet_name}...")
        full_audit[sheet_name] = analyze_sheet(wb[sheet_name])
    
    # Guardamos todo en un JSON temporal para que yo pueda procesarlo sin pérdida
    with open('scratch/full_excel_audit.json', 'w', encoding='utf-8') as f:
        json.dump(full_audit, f, ensure_ascii=False, indent=2)
        
    print(f"\n✅ Auditoría completa finalizada. Se procesaron {len(full_audit)} hojas.")
    for name, content in full_audit.items():
        print(f"- Hoja '{name}': {len(content)} filas detectadas.")

except Exception as e:
    print(f"Error: {e}")
