import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta\Mesas Extraordinarias de Mayo Ciclo 2026 (1).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['_DatesData']
    
    print(f"--- INVENTARIO COMPLETO DE _DatesData (Total Filas: {ws.max_row}) ---")
    
    # Buscamos encabezados o cambios de sección en toda la hoja
    for row_idx in range(1, ws.max_row + 1):
        row = ws[row_idx]
        # Si la fila tiene algún contenido, la analizamos
        vals = [str(cell.value) if cell.value is not None else "" for cell in row[:12]]
        if any(vals):
            # Imprimimos filas que parecen encabezados o inicios de bloques
            if row_idx == 1 or "Carrera" in str(vals) or "GEOGRAFIA" in str(vals).upper() or "POLITICA" in str(vals).upper():
                print(f"\n[SECCIÓN] Fila {row_idx}: {' | '.join(vals)}")
            
            # También imprimimos una muestra cada 50 filas para ver la evolución
            elif row_idx % 50 == 0:
                print(f"Fila {row_idx}: {' | '.join(vals)}")

except Exception as e:
    print(f"Error: {e}")
