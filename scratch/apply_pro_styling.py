import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def apply_pro_styling():
    file_path = r'C:\Users\alanm\Desktop\wqwe\fwew.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1'] # This is Hoja1
    
    # 1. Column Widths
    for col in range(4, 15): # D to N
        sheet.column_dimensions[get_column_letter(col)].width = 16
        
    # Styles Definition
    font_title = Font(size=16, bold=True)
    font_subtitle = Font(size=10.5, italic=True)
    font_label = Font(size=9.5, color="595959")
    font_value = Font(size=12, bold=True)
    font_countdown = Font(size=11, bold=True)
    
    # Iterate rows to find patterns and style
    for r in range(1, sheet.max_row + 1):
        # Look at column D (Subject/Call headers)
        cell_d = sheet.cell(row=r, column=4)
        val_d = str(cell_d.value).strip() if cell_d.value else ""
        
        # Pattern Detection & Styling
        
        # Case 1: Subject Name (e.g., Pedagogía) - Usually follow a gap
        # We detect if it's bold or has specific patterns. 
        # For simplicity, we check Row 3, 10, 17... (pattern seems 7 rows)
        # But let's be smarter: check for "Año" in the row below
        next_val = str(sheet.cell(row=r+1, column=4).value).strip() if sheet.cell(row=r+1, column=4).value else ""
        
        if "Año" in next_val:
            # THIS IS A TITLE ROW
            cell_d.font = font_title
            sheet.row_dimensions[r].height = 32
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            
            # Subtitle (Año 1 | Taller)
            sub_cell = sheet.cell(row=r+1, column=4)
            sub_cell.font = font_subtitle
            sheet.row_dimensions[r+1].height = 20
            
        # Case 2: Call Header (Primer Llamado / Segundo Llamado)
        if "Llamado" in val_d:
            sheet.row_dimensions[r].height = 22
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            # Countdown style in Col N
            count_cell = sheet.cell(row=r, column=14)
            count_cell.font = font_countdown
            count_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Case 3: Labels (Fecha: / Horario:) in Col J
        cell_j = sheet.cell(row=r, column=10)
        val_j = str(cell_j.value).strip() if cell_j.value else ""
        
        if val_j in ["Fecha:", "Horario:"]:
            sheet.row_dimensions[r].height = 24
            cell_j.font = font_label
            cell_j.alignment = Alignment(horizontal="right", vertical="center")
            
            # Value cell in Col L
            val_cell = sheet.cell(row=r, column=12)
            # Remove those weird spaces!
            if isinstance(val_cell.value, str):
                val_cell.value = val_cell.value.strip()
            
            val_cell.font = font_value
            val_cell.alignment = Alignment(horizontal="center", vertical="center")

    output_path = r'C:\Users\alanm\Desktop\wqwe\fwew_STYLISH.xlsx'
    wb.save(output_path)
    print(f"Professional styling applied to Hoja1. Saved at: {output_path}")

if __name__ == "__main__":
    apply_pro_styling()
