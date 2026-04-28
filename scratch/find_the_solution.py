import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb['_DatesData']
    
    print("\n--- BUSCANDO LA SOLUCIÓN OCULTA EN _DATESDATA ---")
    # Buscamos en las columnas de la derecha (K en adelante)
    for r in range(1, 20):
        found = False
        row_str = f"R{r}: "
        for c in range(11, 40): # Hasta la AN
            val = ws.cell(row=r, column=c).value
            if val:
                row_str += f"[{openpyxl.utils.get_column_letter(c)}]: {val} | "
                found = True
        if found:
            print(row_str)

except Exception as e:
    print(f"Error: {e}")
