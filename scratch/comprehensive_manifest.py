import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

def full_column_audit(ws):
    audit = {}
    max_r = ws.max_row
    max_c = ws.max_column
    
    for c in range(1, max_c + 1):
        col_data = []
        has_content = False
        for r in range(1, max_r + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                has_content = True
                col_data.append((r, str(val)))
        
        if has_content:
            audit[openpyxl.utils.get_column_letter(c)] = {
                "total_entries": len(col_data),
                "first_5": col_data[:5],
                "last_5": col_data[-5:] if len(col_data) > 5 else col_data
            }
    return audit

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    comprehensive_manifest = {}
    
    for sheet_name in wb.sheetnames:
        print(f"Auditoría profunda de: {sheet_name}...")
        comprehensive_manifest[sheet_name] = full_column_audit(wb[sheet_name])
    
    with open('scratch/excel_comprehensive_manifest.json', 'w', encoding='utf-8') as f:
        json.dump(comprehensive_manifest, f, ensure_ascii=False, indent=2)
        
    print("\n✅ MANIFIESTO GENERADO. Analizando resultados...")

except Exception as e:
    print(f"Error: {e}")
