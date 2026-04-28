import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

def final_excel_fix_named_range():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_PRO.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Ensure the Dates sheet exists and has a NAME
    if "_DatesData" not in wb.sheetnames:
        data_sheet = wb.create_sheet("_DatesData")
    else:
        data_sheet = wb["_DatesData"]
        data_sheet.delete_rows(1, data_sheet.max_row)
        
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2027, 12, 31)
    current_date = start_date
    row = 1
    while current_date <= end_date:
        data_sheet.cell(row=row, column=1).value = current_date
        current_date += timedelta(days=1)
        row += 1
    
    # Define a NAMED RANGE (Crucial for Excel cross-sheet validation)
    from openpyxl.workbook.defined_name import DefinedName
    new_range = DefinedName("LISTA_FECHAS", attr_text=f"'_DatesData'!$A$1:$A${row-1}")
    if "LISTA_FECHAS" in wb.defined_names:
        del wb.defined_names["LISTA_FECHAS"]
    wb.defined_names.add(new_range)
    
    # 2. Re-apply validation using the Range Name
    # type="list" is better for "solapa"
    dv = DataValidation(type="list", formula1="=LISTA_FECHAS", allow_blank=True)
    dv.errorTitle = 'Error'
    dv.error = 'Seleccione una fecha de la lista.'
    dv.showInputMessage = True
    dv.prompt = 'Haz clic en la flecha para ver las fechas.'
    
    # Clear old validations to avoid conflicts
    sheet.data_validations.dataValidation = []
    sheet.add_data_validation(dv)
    
    # Apply to top-left of merged cells
    for r in [6, 10, 16]:
        dv.add(sheet.cell(row=r, column=15)) # O6, O10, O16
        
    data_sheet.sheet_state = 'hidden'
    
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_ULTIMATE.xlsx'
    wb.save(output_path)
    print(f"ULTIMATE Excel saved at: {output_path}")

if __name__ == "__main__":
    final_excel_fix_named_range()
