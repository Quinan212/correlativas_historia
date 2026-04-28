import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, time, timedelta

def create_chips_style_excel():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_SHEETS_READY.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Update/Clean validations
    sheet.data_validations.dataValidation = []
    
    # Dates Validation
    dv_date = DataValidation(type="list", formula1="=LISTA_FECHAS", allow_blank=True, showDropDown=False) # showDropDown=False means DON'T SUPPRESS (show it)
    sheet.add_data_validation(dv_date)
    
    # Times Validation
    dv_time = DataValidation(type="list", formula1="=LISTA_HORARIOS", allow_blank=True, showDropDown=False)
    sheet.add_data_validation(dv_time)
    
    # 2. Re-apply to merged cells
    for r in range(1, sheet.max_row + 1):
        cell_m = sheet.cell(row=r, column=13) # Column M
        val_m = str(cell_m.value).strip() if cell_m.value else ""
        
        if val_m == "Fecha:":
            dv_date.add(sheet.cell(row=r, column=15)) # O6, O10, O16
        elif val_m == "Horario:":
            dv_time.add(sheet.cell(row=r, column=15)) # O7, O11, O17

    # 3. Final touch: In Google Sheets, this will look like a dropdown.
    # In Excel, we ensure the arrow is not suppressed.
    
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_PARA_SHEETS.xlsx'
    wb.save(output_path)
    print(f"Excel optimized for Sheets with selectors saved at: {output_path}")

if __name__ == "__main__":
    create_chips_style_excel()
