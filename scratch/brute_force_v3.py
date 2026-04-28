import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# VERSION (3) - BRUTE FORCE
file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (3).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Historia']
    
    print("\n--- DUMP TOTAL DE CELDAS CON TEXTO (VERSION 3) ---")
    
    for r in range(1, 400):
        found_in_row = []
        for c in range(1, 40): # Expandimos a columna AN
            val = ws.cell(row=r, column=c).value
            if val:
                found_in_row.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val}")
        
        if found_in_row:
            print(" | ".join(found_in_row))

except Exception as e:
    print(f"Error: {e}")
