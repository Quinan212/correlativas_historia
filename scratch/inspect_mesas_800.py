"""
Inspecciona el Excel "Copia de mesas nuevo 800 VECES MEJOR.xlsx" del escritorio
para entender su estructura antes de transformarlo en un dashboard estático
estilo web-app.
"""
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\alanm\Desktop\Copia de mesas nuevo 800 VECES MEJOR.xlsx"

assert os.path.exists(SRC), f"No existe: {SRC}"

wb = load_workbook(SRC, data_only=False)

print(f"=== Archivo: {SRC} ===")
print(f"Tamaño: {os.path.getsize(SRC):,} bytes")
print(f"Hojas ({len(wb.sheetnames)}): {wb.sheetnames}")
print()

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    print(f"--- Hoja: '{sh_name}' ---")
    print(f"  Dimensiones: {ws.dimensions}  "
          f"(max_row={ws.max_row}, max_col={ws.max_column})")
    print(f"  Freeze: {ws.freeze_panes}   ShowGridLines: {ws.sheet_view.showGridLines}")
    print(f"  TabColor: {ws.sheet_properties.tabColor}")
    # Anchos de columna
    widths = []
    for col in range(1, min(ws.max_column, 30) + 1):
        letter = get_column_letter(col)
        dim = ws.column_dimensions.get(letter)
        w = dim.width if dim else None
        widths.append(f"{letter}={w}")
    print(f"  Anchos (primeras 30): {', '.join(widths)}")
    # Primeras filas de contenido
    print("  Primeras 12 filas × 12 cols:")
    for r in range(1, min(ws.max_row, 12) + 1):
        row_vals = []
        for c in range(1, min(ws.max_column, 12) + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("·")
            else:
                s = str(v).replace("\n", " / ")
                if len(s) > 22:
                    s = s[:22] + "…"
                row_vals.append(s)
        print(f"    R{r:02d}: " + " | ".join(row_vals))
    # Rangos combinados
    merged = list(ws.merged_cells.ranges)
    print(f"  Rangos combinados: {len(merged)}")
    for m in merged[:10]:
        print(f"    - {m}")
    # Imágenes / dibujos
    imgs = getattr(ws, "_images", [])
    print(f"  Imágenes: {len(imgs)}")
    print()
