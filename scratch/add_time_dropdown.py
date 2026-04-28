import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, time, timedelta

def add_time_dropdown():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_ULTIMATE.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Access or create the hidden data sheet
    if "_DatesData" not in wb.sheetnames:
        data_sheet = wb.create_sheet("_DatesData")
    else:
        data_sheet = wb["_DatesData"]
        
    # Generate times from 07:00 to 22:00 every 30 mins in Column B
    current_time = datetime.combine(datetime.today(), time(7, 0))
    end_time = datetime.combine(datetime.today(), time(22, 0))
    row = 1
    while current_time <= end_time:
        data_sheet.cell(row=row, column=2).value = current_time.time()
        data_sheet.cell(row=row, column=2).number_format = 'HH:mm'
        current_time += timedelta(minutes=30)
        row += 1
        
    # Define Named Range for Times
    from openpyxl.workbook.defined_name import DefinedName
    time_range_name = "LISTA_HORARIOS"
    new_range = DefinedName(time_range_name, attr_text=f"'_DatesData'!$B$1:$B${row-1}")
    if time_range_name in wb.defined_names:
        del wb.defined_names[time_range_name]
    wb.defined_names.add(new_range)
    
    # 2. Create Time Validation
    dv_time = DataValidation(type="list", formula1=f"={time_range_name}", allow_blank=True)
    dv_time.errorTitle = 'Horario inválido'
    dv_time.error = 'Selecciona un horario de la lista.'
    dv_time.promptTitle = 'Elegir Horario'
    dv_time.prompt = 'Usa la solapa para seleccionar la hora del examen.'
    
    sheet.add_data_validation(dv_time)
    
    # 3. Identify Time cells (Column O, where Col M is "Horario:")
    for r in range(1, sheet.max_row + 1):
        cell_m = sheet.cell(row=r, column=13) # Column M
        if cell_m.value == "Horario:":
            time_cell = sheet.cell(row=r, column=15) # Column O
            
            # Apply formatting
            time_cell.number_format = 'HH:mm'
            
            # Add to dropdown validation
            dv_time.add(time_cell)
            print(f"Added time solapa to O{r}")
            
    # Save the updated file
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_SHEETS_READY.xlsx'
    wb.save(output_path)
    print(f"Excel ready for Google Sheets saved at: {output_path}")

if __name__ == "__main__":
    add_time_dropdown()
