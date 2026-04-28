import openpyxl
from openpyxl.utils import get_column_letter

file_path = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta\Mesas Extraordinarias de Mayo Ciclo 2026 (1).xlsx'

def get_hex_color(color):
    if color.type == 'rgb':
        return color.rgb
    return "None"

try:
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in ['Mesas Trivunal', 'Coloquios']:
        ws = wb[sheet_name]
        print(f"\n--- ANÁLISIS TÉCNICO DE DISEÑO: {sheet_name} ---")
        
        # Analizar alturas de filas clave
        print("Alturas de fila (Breathing Room):")
        for i in range(5, 15):
            h = ws.row_dimensions[i].height
            print(f"Fila {i}: {h if h else 'Default (15)'} pt")
            
        # Analizar anchos de columnas
        print("\nAnchos de columna:")
        for col in range(1, 6):
            w = ws.column_dimensions[get_column_letter(col)].width
            print(f"Col {get_column_letter(col)}: {w if w else 'Default'} units")

        # Analizar Patrón de Celdas Combinadas
        print("\nPatrón de combinación de celdas (Estructura de Bloques):")
        merged = list(ws.merged_cells.ranges)
        # Mostrar los que están en la zona de la primera materia
        for m in merged[:15]:
            print(f"Combinación: {m}")

        # Analizar Colores y fuentes de un bloque testigo
        test_cell = ws['C6'] # Título de materia
        print(f"\nTestigo Celda C6 ({test_cell.value}):")
        print(f"Color de fondo: {get_hex_color(test_cell.fill.start_color)}")
        print(f"Fuente: {test_cell.font.name}, Tamaño: {test_cell.font.sz}, Bold: {test_cell.font.b}")

except Exception as e:
    print(f"Error: {e}")
