import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

def fix_solapa_merged():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_inteligente.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Re-create the hidden sheet with dates
    if "_DatesData" in wb.sheetnames:
        del wb["_DatesData"]
    data_sheet = wb.create_sheet("_DatesData")
    
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2027, 12, 31)
    current_date = start_date
    row = 1
    while current_date <= end_date:
        data_sheet.cell(row=row, column=1).value = current_date
        data_sheet.cell(row=row, column=1).number_format = 'DD-MM-YYYY'
        current_date += timedelta(days=1)
        row += 1
    data_sheet.sheet_state = 'hidden'
    
    # Create Data Validation
    date_range = f"=_DatesData!$A$1:$A${row-1}"
    dv = DataValidation(type="list", formula1=date_range, allow_blank=True)
    
    # Custom error/prompt
    dv.errorTitle = 'Selección de Fecha'
    dv.error = 'Por favor selecciona una fecha de la lista desplegable.'
    dv.promptTitle = 'Selector de Fecha'
    dv.prompt = 'Haz clic en la flecha para elegir la fecha del examen.'
    
    sheet.add_data_validation(dv)
    
    # Highlight fill (subtle gold/yellow to indicate it's a field)
    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Apply to the EXACT merged ranges found previously
    # Row 6: O6:P6
    # Row 10: O10:P10
    # Row 16: O16:P16
    date_cells = ["O6", "O10", "O16"]
    
    for coord in date_cells:
        cell = sheet[coord]
        dv.add(cell)
        cell.fill = fill
        print(f"Fixed solapa for merged cell {coord}")

    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_PRO.xlsx'
    wb.save(output_path)
    print(f"Final Excel saved at: {output_path}")

if __name__ == "__main__":
    fix_solapa_merged()
