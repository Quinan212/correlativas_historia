import openpyxl

def check_structure():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_con_solapa.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    print("--- Merged Cells ---")
    for merged_range in sheet.merged_cells.ranges:
        print(merged_range)
        
    print("\n--- Content around expected date cells ---")
    for row in range(1, 20):
        m_val = sheet.cell(row=row, column=13).value
        o_val = sheet.cell(row=row, column=15).value
        if m_val == "Fecha:":
            print(f"Row {row}: Col M='{m_val}', Col O='{o_val}'")

if __name__ == "__main__":
    check_structure()
