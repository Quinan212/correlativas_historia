import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Coloquios']
    
    print("\n--- EL BLOQUE DE COLOQUIOS (Pedagogía 1º Año) ---")
    for r in range(6, 16):
        line = []
        for c in range(1, 20):
            val = ws.cell(row=r, column=c).value
            if val:
                line.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val}")
        if line:
            print(" | ".join(line))

except Exception as e:
    print(f"Error: {e}")
