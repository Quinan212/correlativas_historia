import openpyxl
from openpyxl.styles import PatternFill

def unmerge_and_fix_for_sheets():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_COMPATIBLE_SHEETS.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Identify and UNMERGE specifically the Date and Time cells
    # We know from check_structure they were O6:P6, O10:P10, O16:P16, etc.
    # We will unmerge all ranges that start with O in the target rows
    
    target_rows = []
    for r in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row=r, column=13).value).strip() if sheet.cell(row=r, column=13).value else ""
        if label in ["Fecha:", "Horario:"]:
            target_rows.append(r)
            
    # List of merged ranges to remove
    to_unmerge = []
    for merged_range in sheet.merged_cells.ranges:
        # Check if the range starts at Column O (15) and is in our target rows
        if merged_range.min_col == 15 and merged_range.min_row in target_rows:
            to_unmerge.append(merged_range)
            
    for m_range in to_unmerge:
        sheet.unmerge_cells(str(m_range))
        print(f"Unmerged {m_range} to ensure Sheets compatibility.")

    # 2. Make sure the background color is only on Column O now
    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for r in target_rows:
        sheet.cell(row=r, column=15).fill = fill
        # Clear P column formatting if needed to avoid phantom borders
        sheet.cell(row=r, column=16).fill = PatternFill(fill_type=None)

    # 3. Save as the final version for Sheets
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_DIRECTO_A_SHEETS.xlsx'
    wb.save(output_path)
    print(f"Final unmerged Excel saved at: {output_path}")

if __name__ == "__main__":
    unmerge_and_fix_for_sheets()
