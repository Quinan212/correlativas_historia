import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

def analyze_sheet_complete(ws):
    data = []
    # Usamos el rango completo detectado por Excel
    for row in ws.iter_rows(values_only=True):
        # Capturamos todas las filas, incluso las que parecen vacías pero tienen formato
        data.append([str(c) if c is not None else "" for c in row])
    return data

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    full_audit_v2 = {}
    
    for sheet_name in wb.sheetnames:
        print(f"Leyendo hoja: {sheet_name}...")
        full_audit_v2[sheet_name] = analyze_sheet_complete(wb[sheet_name])
    
    with open('scratch/full_excel_audit_v2.json', 'w', encoding='utf-8') as f:
        json.dump(full_audit_v2, f, ensure_ascii=False, indent=2)
        
    print(f"\n✅ LECTURA TOTAL FINALIZADA DE LA VERSIÓN 2.")
    for name, content in full_audit_v2.items():
        print(f"- Hoja '{name}': {len(content)} filas y {len(content[0]) if content else 0} columnas procesadas.")

except Exception as e:
    print(f"Error: {e}")
