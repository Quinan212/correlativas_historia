import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    
    for sn in ['Mesas Trivunal', 'Coloquios']:
        ws = wb[sn]
        print(f"\n--- COLUMNAS A y B DE {sn} ---")
        for r in range(1, 150):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value
            if a_val or b_val:
                print(f"R{r}: [A]: {a_val} | [B]: {b_val}")

except Exception as e:
    print(f"Error: {e}")
