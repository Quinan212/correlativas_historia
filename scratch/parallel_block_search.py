import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Mesas Trivunal']
    
    print("\n--- MATERIAS ENCONTRADAS POR COLUMNA ---")
    # Escaneamos muchas columnas (hasta la ZZZ)
    for c in range(1, 100):
        col_letter = openpyxl.utils.get_column_letter(c)
        found_in_col = []
        for r in range(1, 200):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and len(val.strip()) > 15:
                if "AÑO" not in val and "Acta" not in val:
                    found_in_col.append(f"[Fila {r}]: {val.strip()}")
        
        if found_in_col:
            print(f"\nCOLUMNA {col_letter}:")
            for entry in found_in_col:
                print(f"  {entry}")

except Exception as e:
    print(f"Error: {e}")
