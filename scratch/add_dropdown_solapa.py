import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

def add_dropdown_solapa():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_inteligente.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Create a "Data" sheet for the dates if it doesn't exist
    if "_DatesData" in wb.sheetnames:
        del wb["_DatesData"]
    data_sheet = wb.create_sheet("_DatesData")
    
    # Generate dates for 2026 and 2027
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2027, 12, 31)
    current_date = start_date
    row = 1
    while current_date <= end_date:
        data_sheet.cell(row=row, column=1).value = current_date
        data_sheet.cell(row=row, column=1).number_format = 'DD-MM-YYYY'
        current_date += timedelta(days=1)
        row += 1
    
    # Hide the data sheet
    data_sheet.sheet_state = 'hidden'
    
    # 2. Define the validation range (pointing to the hidden sheet)
    # The range is _DatesData!$A$1:$A$730 (approx 2 years)
    date_range = f"=_DatesData!$A$1:$A${row-1}"
    
    dv = DataValidation(type="list", formula1=date_range, showDropDown=False) # showDropDown=False in code logic actually means 'the drop down arrow is shown' in some contexts, but let's check. 
    # Actually, showDropDown=False in openpyxl means 'suppress the arrow', which we DON'T want.
    # By default, openpyxl DataValidation list shows the arrow.
    
    dv.errorTitle = 'Selección inválida'
    dv.error = 'Por favor, selecciona una fecha de la lista.'
    dv.promptTitle = 'Elegir Fecha'
    dv.prompt = 'Usa la solapa para elegir la fecha del examen.'
    
    sheet.add_data_validation(dv)
    
    # 3. Apply to cells
    for r in range(1, sheet.max_row + 1):
        cell_m = sheet.cell(row=r, column=13) # Column M
        if cell_m.value == "Fecha:":
            date_cell = sheet.cell(row=r, column=15) # Column O
            dv.add(date_cell)
            print(f"Added dropdown solapa to O{r}")
            
    # Save the final version
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_con_solapa.xlsx'
    wb.save(output_path)
    print(f"Excel with solapa saved at: {output_path}")

if __name__ == "__main__":
    add_dropdown_solapa()
