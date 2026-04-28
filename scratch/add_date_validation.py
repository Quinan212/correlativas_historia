import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation

def automate_and_format_excel():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_automatizado.xlsx'
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Create a nice date style if it doesn't exist
    date_style_name = "CustomDateStyle"
    if date_style_name not in wb.style_names:
        date_style = NamedStyle(name=date_style_name)
        date_style.number_format = 'DD-MM-YYYY' # Or 'DD "de" MMMM "de" YYYY'
        wb.add_named_style(date_style)

    # Data Validation for dates (allows any date but ensures it is a date)
    dv = DataValidation(type="date", operator="greaterThan", formula1="1/1/1900")
    dv.errorTitle = 'Fecha inválida'
    dv.error = 'Por favor, ingresa una fecha válida.'
    dv.promptTitle = 'Ingresar Fecha'
    dv.prompt = 'Escribe la fecha del examen (ej: 05/05/2026)'
    
    sheet.add_data_validation(dv)
    
    for row in range(1, sheet.max_row + 1):
        cell_m = sheet.cell(row=row, column=13) # Column M
        if cell_m.value == "Fecha:":
            date_cell = sheet.cell(row=row, column=15) # Column O
            
            # Apply the style
            date_cell.style = date_style_name
            
            # Center the text
            date_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add to data validation range
            dv.add(date_cell)
            
            print(f"Applied date formatting and validation to O{row}")

    # Save changes
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_inteligente.xlsx'
    wb.save(output_path)
    print(f"Intelligent Excel saved at: {output_path}")

if __name__ == "__main__":
    automate_and_format_excel()
