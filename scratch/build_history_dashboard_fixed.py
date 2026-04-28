import openpyxl
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_history_dashboard():
    # Base template (using the one with dropdown data ready)
    template_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_COMPATIBLE_SHEETS.xlsx'
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 1. Define the subjects list
    mesas_mayo = [
        # 1st Year
        ("Problemática del Conocimiento Histórico", "1° Año | Asignatura"),
        ("Pedagogía", "1° Año | Asignatura"),
        ("Didáctica General", "1° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales de la Antigüedad", "1° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales de los Pueblos Originarios de América", "1° Año | Asignatura"),
        # 2nd Year
        ("Procesos Políticos, Sociales, Económicos y Culturales del Feudalismo y la Modernidad", "2° Año | Asignatura"),
        ("Psicología Educacional", "2° Año | Asignatura"),
        ("Procesos Americanos I", "2° Año | Asignatura"),
        ("Filosofía", "2° Año | Asignatura"),
        # 3rd Year
        ("Historia de la Educación Argentina", "3° Año | Asignatura"),
        ("Didáctica de la Historia", "3° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Americanos II", "3° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Contemporáneos I", "3° Año | Asignatura"),
        ("Sociología de la Educación", "3° Año | Asignatura"),
        # 4th Year
        ("Procesos Sociales, Políticos, Económicos y Culturales Contemporáneos II", "4° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Americanos III", "4° Año | Asignatura"),
        ("Derechos Humanos", "4° Año | Asignatura"),
    ]
    
    coloquios = [
        ("Sujeto de la Educación Secundaria", "Coloquio | Taller"),
        ("Práctica Docente I", "Coloquio | Taller"),
        ("Práctica Docente II", "Coloquio | Taller"),
        ("Práctica Docente III", "Coloquio | Taller"),
        ("Epistemología de la Historia", "Coloquio | Seminario"),
        ("Didáctica de las Ciencias Sociales", "Coloquio | Seminario"),
    ]

    # Clean existing data below row 1
    # Note: we need to keep the hidden columns Z and AA if they are there
    sheet.delete_rows(2, 500)
    
    # Re-apply Data Validations (Google Sheets style)
    # Range for Dates: $Z$1:$Z$730
    # Range for Times: $AA$1:$AA$31
    dv_date = DataValidation(type="list", formula1="=$Z$1:$Z$730", allow_blank=True)
    dv_time = DataValidation(type="list", formula1="=$AA$1:$AA$31", allow_blank=True)
    sheet.add_data_validation(dv_date)
    sheet.add_data_validation(dv_time)

    def write_styled_block(subject, info, call, r):
        # Basic Premium Styles
        font_name = Font(size=12, bold=True)
        font_info = Font(size=10, italic=True)
        fill_header = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        fill_call = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        fill_data = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Row 1: Name
        sheet.cell(row=r, column=6).value = subject
        sheet.merge_cells(start_row=r, start_column=6, end_row=r, end_column=17)
        sheet.cell(row=r, column=6).font = font_name
        sheet.cell(row=r, column=6).fill = fill_header
        sheet.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        
        # Row 2: Info
        sheet.cell(row=r+1, column=6).value = info
        sheet.merge_cells(start_row=r+1, start_column=6, end_row=r+1, end_column=17)
        sheet.cell(row=r+1, column=6).font = font_info
        sheet.cell(row=r+1, column=6).alignment = Alignment(horizontal="center")
        
        # Row 3: Call + Countdown
        sheet.cell(row=r+2, column=6).value = call
        sheet.merge_cells(start_row=r+2, start_column=6, end_row=r+2, end_column=15)
        sheet.cell(row=r+2, column=6).fill = fill_call
        sheet.cell(row=r+2, column=6).alignment = Alignment(horizontal="center")
        
        countdown_cell = sheet.cell(row=r+2, column=16)
        sheet.merge_cells(start_row=r+2, start_column=16, end_row=r+2, end_column=17)
        date_ref = f"O{r+3}"
        countdown_cell.value = f'=IF({date_ref}-TODAY()=0, "Hoy", IF({date_ref}-TODAY()=1, "Mañana", IF({date_ref}-TODAY()<0, "Finalizado", "En " & ({date_ref}-TODAY()) & " días")))'
        countdown_cell.alignment = Alignment(horizontal="center")
        countdown_cell.font = Font(bold=True)
        
        # Row 4: Fecha
        sheet.cell(row=r+3, column=12).value = "Fecha:"
        date_cell = sheet.cell(row=r+3, column=15)
        sheet.merge_cells(start_row=r+3, start_column=15, end_row=r+3, end_column=17)
        date_cell.fill = fill_data
        date_cell.number_format = 'DD-MM-YYYY'
        date_cell.alignment = Alignment(horizontal="center")
        dv_date.add(date_cell)
        
        # Row 5: Horario
        sheet.cell(row=r+4, column=12).value = "Horario:"
        time_cell = sheet.cell(row=r+4, column=15)
        sheet.merge_cells(start_row=r+4, start_column=15, end_row=r+4, end_column=17)
        time_cell.fill = fill_data
        time_cell.number_format = 'HH:mm'
        time_cell.alignment = Alignment(horizontal="center")
        dv_time.add(time_cell)
        
        return r + 7

    # 2. Build Banners and Blocks
    row = 2
    # Banner Mayo
    banner_font = Font(size=14, bold=True, color="FFFFFF")
    banner_fill_mayo = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sheet.cell(row=row, column=6).value = "📅 EXÁMENES FINAL - TURNO MAYO 2026"
    sheet.merge_cells(f"F{row}:Q{row}")
    sheet.cell(row=row, column=6).font = banner_font
    sheet.cell(row=row, column=6).fill = banner_fill_mayo
    sheet.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    
    row += 2
    
    # ADD MAYO SUBJECTS
    for s, info in mesas_mayo:
        row = write_styled_block(s, info, "Llamado de Mayo", row)
        
    # Banner Coloquios
    row += 1
    sheet.cell(row=row, column=6).value = "🎓 COLOQUIOS / TALLERES - HISTORIA 2026"
    sheet.merge_cells(f"F{row}:Q{row}")
    banner_fill_coloquio = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    sheet.cell(row=row, column=6).font = banner_font
    sheet.cell(row=row, column=6).fill = banner_fill_coloquio
    sheet.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    row += 2
    
    # Build Coloquios
    for s, info in coloquios:
        row = write_styled_block(s, info, "Coloquio", row)

    output_path = r'C:\Users\alanm\Desktop\wqwe\Tablero_Historia_Mayo_2026.xlsx'
    wb.save(output_path)
    print(f"Full Dashboard saved at: {output_path}")

if __name__ == "__main__":
    create_history_dashboard()
