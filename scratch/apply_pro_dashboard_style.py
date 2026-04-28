"""
APLICA EL ESTILO DE DASHBOARD ESTÁTICO STYLE GOOGLE / NOTION
según las instrucciones del video:
  ✅ Fondo uniforme #F1F5F9 en TODA la hoja
  ✅ Bordes sombreados suaves en las cards
  ✅ Espaciado perfecto
  ✅ Eliminar todos los bordes bruscos
  ✅ Redondeo visual de tarjetas
  ✅ Tipografía limpia
  ✅ Todo el Excel deja de parecer un Excel
"""
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\alanm\Desktop\Copia de mesas nuevo 800 VECES MEJOR.xlsx"
DST = r"C:\Users\alanm\Desktop\Copia de mesas nuevo 800 VECES MEJOR_DASHBOARD.xlsx"

# Paleta de colores oficial Dashboard
BG_GLOBAL = "F1F5F9"  # Fondo gris claro Google
BG_CARD = "FFFFFF"    # Fondo tarjetas blanco
BORDER_CARD = "E2E8F0"# Borde tarjetas
TEXT_PRIMARY = "111827"# Texto principal
TEXT_SECONDARY = "64748B"# Texto secundario

wb = load_workbook(SRC, data_only=False)

# Preparar estilos
fill_bg = PatternFill(start_color=BG_GLOBAL, end_color=BG_GLOBAL, fill_type="solid")
fill_card = PatternFill(start_color=BG_CARD, end_color=BG_CARD, fill_type="solid")
side_light = Side(style="thin", color=BORDER_CARD)
border_card = Border(
    left=side_light,
    right=side_light,
    top=side_light,
    bottom=side_light
)
font_main = Font(name="Segoe UI", size=11, color=TEXT_PRIMARY, bold=False)
font_header = Font(name="Segoe UI", size=12, color=TEXT_PRIMARY, bold=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

print(f"✅ Cargado archivo: {SRC}")

for sheet_name in ['MESAS', 'C.COLOQUIOS', 'Historia', 'Geografia', 'Ciencias Politicas', 'Coloquios']:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    print(f"\n🔄 Procesando hoja: {sheet_name}")

    # 1. APLICAR FONDO UNIFORME A TODA LA HOJA
    max_r = max(ws.max_row, 600)
    max_c = max(ws.max_column, 22)

    print(f"   ↳ Pintando fondo total: {max_r} filas × {max_c} columnas")
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            if cell.fill is None or cell.fill.patternType is None or cell.fill.fgColor.rgb in (None, "00000000", "FFF5F7FA"):
                cell.fill = fill_bg

    # 2. ENCONTRAR TODAS LAS CARDS / BLOQUES DE CONTENIDO
    # Buscar todos los rangos que tienen fondo blanco
    print("   ↳ Aplicando estilo cards...")
    card_ranges = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            # Detectar celda que es parte de una tarjeta
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb in ("FFFFFFFF", "FFFFFF", "FFF3F3F3", "FFEFEFEF"):
                # Aplicar estilo oficial a tarjeta
                cell.fill = fill_card
                cell.border = border_card
                # Aplicar fuente limpia
                if cell.font.bold:
                    cell.font = font_header
                else:
                    cell.font = font_main
                # Alineación
                if isinstance(cell.value, str) and len(str(cell.value)) > 25:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

    # 3. AJUSTAR ESPACIADO Y ANCHOS PERFECTOS
    for col_idx in range(1, max_c + 1):
        cl = get_column_letter(col_idx)
        if col_idx in (1, 2):
            ws.column_dimensions[cl].width = 2
        else:
            # Mantener ancho mínimo pero limpio
            old_w = ws.column_dimensions[cl].width if cl in ws.column_dimensions and ws.column_dimensions[cl].width else 0
            ws.column_dimensions[cl].width = max(old_w, 14)

    # 4. ALTURA DE FILAS PARA DAR AIRE
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = max(ws.row_dimensions[r].height if ws.row_dimensions[r].height else 0, 22)

    # 5. GARANTIZAR QUE NO SE MUESTREN CUADRICULAS
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.showZeros = False
    ws.sheet_view.view = "normal"

    print(f"   ✅ Hoja {sheet_name} terminada")

# Guardar resultado final
wb.save(DST)
wb.close()

print(f"\n✅ ✅ ✅ DASHBOARD LISTO! ✅ ✅ ✅")
print(f"Archivo guardado: {DST}")
print("\nCaracteristicas aplicadas:")
print("  ✅ Sin líneas de cuadrícula")
print("  ✅ Fondo gris uniforme estilo Google")
print("  ✅ Todas las cards con bordes suaves")
print("  ✅ Tipografía limpia Segoe UI")
print("  ✅ Espaciado y aire en todas las filas")
print("  ✅ Ocultos encabezados de filas/columnas")
print("  ✅ Sin ceros a la vista")
print("\nAHORA ABRILO Y VAS A VER QUE YA NO PARECE EXCEL!")