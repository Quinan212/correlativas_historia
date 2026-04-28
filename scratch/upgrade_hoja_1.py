import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def upgrade_hoja_1():
    file_path = r'C:\Users\alanm\Desktop\wqwe\fwew (1).xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Hoja 1']
    
    # Colors
    color_banner = "2E75B6" # Blue
    color_subject = "D9EAD3" # Light Green
    color_label = "595959" # Dark Gray
    color_division = "4472C4" # Blueish Gray
    
    # Font Styles
    font_main_banner = Font(size=22, bold=True, color="FFFFFF")
    font_subject = Font(size=18, bold=True)
    font_division = Font(size=11, bold=True, color="FFFFFF")
    font_label = Font(size=10, color=color_label)
    font_value = Font(size=12, bold=True)
    font_teacher = Font(size=11, color="333333")
    
    # 1. MAIN BANNER (Coloquios) - Row 2
    sheet.row_dimensions[2].height = 50
    # Merge if not merged
    sheet.merge_cells("A2:N2")
    banner_cell = sheet.cell(row=2, column=1)
    banner_cell.value = "📋 PANEL DE COLOQUIOS - HISTORIA"
    banner_cell.font = font_main_banner
    banner_cell.fill = PatternFill(start_color=color_banner, end_color=color_banner, fill_type="solid")
    banner_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Iterate through rows to find patterns
    for r in range(3, sheet.max_row + 1):
        # Detect Subject Name (Row 6 had it in Col B in my previous read)
        cell_b = sheet.cell(row=r, column=2)
        val_b = str(cell_b.value).strip() if cell_b.value else ""
        
        # Look for the subject header (large text block)
        if "Práctica Docente" in val_b or (len(val_b) > 10 and r > 5 and sheet.cell(row=r+1, column=2).value is not None):
             sheet.row_dimensions[r].height = 40
             cell_b.font = font_subject
             cell_b.alignment = Alignment(horizontal="center", vertical="center")
             # Merge it horizontally for impact
             try:
                 sheet.merge_cells(start_row=r, start_column=2, end_row=r, end_column=14)
             except: pass
             sheet.cell(row=r, column=2).fill = PatternFill(start_color=color_subject, end_color=color_subject, fill_type="solid")

        # Detect Division Headers (División A / B)
        cell_c = sheet.cell(row=r, column=3)
        val_c = str(cell_c.value).strip() if cell_c.value else ""
        if "División" in val_b or "División" in val_c:
            sheet.row_dimensions[r].height = 25
            style_cell = cell_b if "División" in val_b else cell_c
            style_cell.font = font_division
            style_cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            style_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Detect Labels (Fecha: / Horario:) in Col J
        cell_j = sheet.cell(row=r, column=10)
        val_j = str(cell_j.value).strip() if cell_j.value else ""
        if val_j in ["Fecha:", "Horario:"]:
            sheet.row_dimensions[r].height = 28
            cell_j.font = font_label
            cell_j.alignment = Alignment(horizontal="right", vertical="center")
            
            # Value cell in Col L
            val_cell = sheet.cell(row=r, column=12)
            if isinstance(val_cell.value, str):
                val_cell.value = val_cell.value.replace("---", "").strip()
            val_cell.font = font_value
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Cleanup "---" in names (Col D, G)
        for col_idx in [4, 7]:
            cell = sheet.cell(row=r, column=col_idx)
            if str(cell.value).strip() == "---------------":
                cell.value = ""
            elif cell.value:
                cell.font = font_teacher

    # Hide gridlines for a cleaner look
    sheet.sheet_view.showGridLines = False
    
    output_path = r'C:\Users\alanm\Desktop\wqwe\fwew_1_DASHBOARD_PRO.xlsx'
    wb.save(output_path)
    print(f"Hoja 1 fully upgraded. Saved at: {output_path}")

if __name__ == "__main__":
    upgrade_hoja_1()
