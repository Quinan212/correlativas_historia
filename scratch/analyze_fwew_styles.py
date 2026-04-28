import openpyxl

def analyze_styles():
    file_path = r'C:\Users\alanm\Desktop\wqwe\fwew.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Analyze the first block (Rows 3 to 7)
    for r in range(3, 8):
        print(f"--- Row {r} (Height: {sheet.row_dimensions[r].height}) ---")
        for c in range(4, 15): # Col D to N
            cell = sheet.cell(row=r, column=c)
            # Only columns with content or interesting merges
            if cell.value or any(r in range(m.min_row, m.max_row + 1) and c in range(m.min_col, m.max_col + 1) for m in sheet.merged_cells.ranges):
                font = cell.font
                align = cell.alignment
                print(f"Col {c} ({sheet.column_dimensions[openpyxl.utils.get_column_letter(c)].width}): Value='{cell.value}', Font={font.sz}, Bold={font.b}, Align={align.horizontal}")

if __name__ == "__main__":
    analyze_styles()
