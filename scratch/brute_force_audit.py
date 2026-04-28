import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n===== HOJA: {sheet_name} =====")
        
        # Encontramos los límites reales buscando la última celda con contenido
        real_max_row = 0
        real_max_col = 0
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    real_max_row = max(real_max_row, cell.row)
                    real_max_col = max(real_max_col, cell.column)
        
        print(f"Límite real detectado: Fila {real_max_row}, Columna {real_max_col}")
        
        # Ahora listamos los bloques de datos detectados
        for r in range(1, real_max_row + 1):
            row_data = []
            for c in range(1, real_max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_data.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val}")
            
            if row_data:
                # Solo imprimimos si la fila tiene algo, para no saturar
                print(" | ".join(row_data))

except Exception as e:
    print(f"Error: {e}")
