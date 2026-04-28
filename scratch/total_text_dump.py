import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sn in ['Mesas Trivunal', 'Coloquios']:
        ws = wb[sn]
        print(f"\n===== VOLCADO TOTAL DE TEXTOS: {sn} =====")
        for r in range(1, 600):
            for c in range(1, 150):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    t = val.strip().replace('\n', ' ')
                    if len(t) > 2:
                        print(f"{sn} [{openpyxl.utils.get_column_letter(c)}{r}]: {t}")

except Exception as e:
    print(f"Error: {e}")
