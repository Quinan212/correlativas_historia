import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Mesas Trivunal']
    
    # Rango extendido para encontrar bloques paralelos
    for r in range(1, 400):
        row_data = []
        for c in range(1, 150): # Escaneamos hasta la columna ET
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and len(val.strip()) > 3:
                row_data.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val.strip()}")
        
        if row_data:
            print(" | ".join(row_data))

except Exception as e:
    print(f"Error: {e}")
