import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# AHORA SI: VERSION (3)
file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (3).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Historia'] # Asumimos que la hoja se llama Historia ahora
    
    print("\n--- MATERIAS ENCONTRADAS EN LA VERSION (3) ---")
    
    found_count = 0
    for r in range(1, 400):
        for c in [4, 8]: # Columnas D y H
            val = ws.cell(row=r, column=c).value
            if val:
                val_str = str(val).strip()
                # Un filtro mucho más relajado: solo ignoramos lo obvio
                if (len(val_str) > 5 and 
                    'Año' not in val_str and 
                    'Medina' not in val_str and 
                    'Igual' not in val_str and 
                    'Borche' not in val_str and 
                    'Frigo' not in val_str and 
                    'Acta' not in val_str and 
                    '---' not in val_str):
                    
                    found_count += 1
                    # Buscamos la fecha en L o M
                    date_cell = "N/A"
                    for dr in range(r, r + 10):
                        for dc in [12, 13]: # L o M
                            dv = ws.cell(row=dr, column=dc).value
                            if dv and ("2026" in str(dv) or (isinstance(dv, (int, float)) and 45000 < dv < 46500)):
                                date_cell = f"{openpyxl.utils.get_column_letter(dc)}{dr}"
                                break
                        if date_cell != "N/A": break
                    
                    print(f"{found_count}. [D{r}]: {val_str} | Fecha: {date_cell}")

except Exception as e:
    # Si falla por nombre de hoja, probamos Mesas Trivunal
    try:
        ws = wb['Mesas Trivunal']
        print("\n--- MATERIAS ENCONTRADAS EN 'Mesas Trivunal' (VERSION 3) ---")
        found_count = 0
        for r in range(1, 400):
            for c in [4]: 
                val = ws.cell(row=r, column=c).value
                if val:
                    val_str = str(val).strip()
                    if (len(val_str) > 5 and 'Año' not in val_str and 'Docente' not in val_str and 'Acta' not in val_str and '---' not in val_str):
                        found_count += 1
                        print(f"{found_count}. [D{r}]: {val_str}")
    except:
        print(f"Error: {e}")
