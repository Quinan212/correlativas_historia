import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

def automate_excel():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq.xlsx'
    
    # Load the workbook preserving styles
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Iterate through rows to find "Fecha:" labels and their values
    # Based on previous analysis:
    # Date label is in column 13 (M), Date value is in column 15 (O)
    # Countdown is in column 17 (Q), one row ABOVE the date row
    
    for row in range(1, sheet.max_row + 1):
        cell_m = sheet.cell(row=row, column=13) # Column M
        if cell_m.value == "Fecha:":
            date_cell_ref = f"O{row}" # Column O
            countdown_row = row - 1
            countdown_cell = sheet.cell(row=countdown_row, column=17) # Column Q
            
            # Formula in Spanish for Excel
            # If (Date - TODAY) == 0 -> "Hoy"
            # If (Date - TODAY) == 1 -> "Mañana"
            # If (Date - TODAY) < 0 -> "Finalizado"
            # Else -> "En X días"
            
            # Note: We use commas or semicolons depending on Excel locale. 
            # Usually English API uses commas, and Excel translates to locale.
            # However, for string concatenation in Spanish Excel it might need adjustment.
            # We'll use the standard Excel formula format.
            
            formula = (
                f'=IF({date_cell_ref}-TODAY()=0, "Hoy", '
                f'IF({date_cell_ref}-TODAY()=1, "Mañana", '
                f'IF({date_cell_ref}-TODAY()<0, "Finalizado", '
                f'"En " & ({date_cell_ref}-TODAY()) & " días")))'
            )
            
            countdown_cell.value = formula
            print(f"Applied formula for {date_cell_ref} at Q{countdown_row}")

    # Save the updated workbook
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_automatizado.xlsx'
    wb.save(output_path)
    print(f"Automated Excel saved at: {output_path}")

if __name__ == "__main__":
    automate_excel()
