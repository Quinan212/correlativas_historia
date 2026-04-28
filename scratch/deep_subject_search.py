import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Escaneamos TODOS los nombres de materias posibles en TODAS las hojas
    for sheet_name in wb.sheetnames:
        print(f"\n--- BUSCANDO MATERIAS EN HOJA: {sheet_name} ---")
        ws = wb[sheet_name]
        for r in range(1, 500):
            for c in range(1, 100):
                val = ws.cell(row=r, column=c).value
                # Buscamos nombres de materias largos que no sean fechas ni encabezados fijos
                if val and isinstance(val, str) and len(val.strip()) > 15:
                    if "AÑO" not in val and "Acta" not in val and "Sin Inscrip" not in val:
                        print(f"[{openpyxl.utils.get_column_letter(c)}{r}]: {val.strip()}")

except Exception as e:
    print(f"Error: {e}")
