import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\alanm\Desktop\wqwe\Mesas Extraordinarias de Mayo Ciclo 2026 (2).xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Mesas Trivunal']
    
    print("\n--- MAPA DEFINITIVO DE TODAS LAS MATERIAS ---")
    
    # Buscamos en todas las columnas posibles
    found_count = 0
    for r in range(1, 400):
        # En tu diseño las materias suelen estar en la columna D o H (las dos principales)
        for c in [4, 8]: 
            val = ws.cell(row=r, column=c).value
            if val:
                val_str = str(val).strip()
                # Filtramos nombres de profes y etiquetas de 'Año'
                if len(val_str) > 5 and 'Año' not in val_str and 'Docente' not in val_str and 'Acta' not in val_str and '---' not in val_str and ',' not in val_str:
                    found_count += 1
                    # IMPORTANTE: Buscamos la fecha en la columna L de la misma fila o un par de filas abajo
                    date_val = None
                    for dr in range(r, r + 5):
                        dv = ws.cell(row=dr, column=12).value # Columna L
                        if dv and ("2026" in str(dv) or (isinstance(dv, (int, float)) and 45000 < dv < 46500)):
                            date_val = f"L{dr}"
                            break
                    
                    if not date_val:
                        date_val = "Verificar L"
                        
                    print(f"{found_count}. Materia [D{r}]: {val_str} | Fecha: {date_val}")

except Exception as e:
    print(f"Error: {e}")
