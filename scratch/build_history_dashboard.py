import openpyxl
from copy import copy

def create_history_dashboard():
    # Base template (using the one with dropdown data ready)
    template_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_MAESTRO_COMBINADO.xlsx'
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 1. Define the subjects list
    mesas_mayo = [
        # 1st Year
        ("Problemática del Conocimiento Histórico", "1° Año | Asignatura"),
        ("Pedagogía", "1° Año | Asignatura"),
        ("Didáctica General", "1° Año | Asignatura"),
        ("Procesos Históricos, Políticos, Económicos y Culturales de la Antigüedad", "1° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales de los Pueblos Originarios de América", "1° Año | Asignatura"),
        # 2nd Year
        ("Procesos Políticos, Sociales, Económicos y Culturales del Feudalismo y la Modernidad", "2° Año | Asignatura"),
        ("Psicología Educacional", "2° Año | Asignatura"),
        ("Procesos Americanos I", "2° Año | Asignatura"),
        ("Filosofía", "2° Año | Asignatura"),
        # 3rd Year
        ("Historia de la Educación Argentina", "3° Año | Asignatura"),
        ("Didáctica de la Historia", "3° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Americanos II", "3° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Contemporáneos I", "3° Año | Asignatura"),
        ("Sociología de la Educación", "3° Año | Asignatura"),
        # 4th Year
        ("Procesos Sociales, Políticos, Económicos y Culturales Contemporáneos II", "4° Año | Asignatura"),
        ("Procesos Sociales, Políticos, Económicos y Culturales Americanos III", "4° Año | Asignatura"),
        ("Derechos Humanos", "4° Año | Asignatura"),
    ]
    
    coloquios = [
        ("Sujeto de la Educación Secundaria", "Coloquio | Taller"),
        ("Práctica Docente I", "Coloquio | Taller"),
        ("Práctica Docente II", "Coloquio | Taller"),
        ("Práctica Docente III", "Coloquio | Taller"),
        ("Epistemología de la Historia", "Coloquio | Seminario"),
        ("Didáctica de las Ciencias Sociales", "Coloquio | Seminario"),
    ]

    # Clear current sheet content except the template block if any
    # Actually we'll just append after the first two examples
    
    current_row = 1
    
    def copy_block(subject_name, subject_info, call_type, start_row):
        # Template block is Rows 2-7
        source_start = 2
        block_size = 6
        
        for i in range(block_size):
            src_row = source_start + i
            dst_row = start_row + i
            
            # Copy values and styles
            for col in range(1, 20):
                source_cell = sheet.cell(row=src_row, column=col)
                target_cell = sheet.cell(row=dst_row, column=col)
                
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = copy(source_cell.number_format)
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

        # Inject specific data
        sheet.cell(row=start_row, column=6).value = subject_name
        sheet.cell(row=start_row+1, column=6).value = subject_info
        sheet.cell(row=start_row+2, column=6).value = call_type
        
        # Inject Date drop down and formulas
        date_cell_ref = f"O{start_row+4}"
        countdown_cell = sheet.cell(row=start_row+2, column=17)
        countdown_cell.value = f'=IF({date_cell_ref}-TODAY()=0, "Hoy", IF({date_cell_ref}-TODAY()=1, "Mañana", IF({date_cell_ref}-TODAY()<0, "Finalizado", "En " & ({date_cell_ref}-TODAY()) & " días")))'
        
        # Copy merged cells structure
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row >= source_start and merged_range.max_row <= (source_start + block_size - 1):
                shift = start_row - source_start
                new_range = f"{merged_range.min_col_base}{merged_range.min_row + shift}:{merged_range.max_col_base}{merged_range.max_row + shift}"
                # Avoid adding existing merges
                try:
                    sheet.merge_cells(new_range)
                except:
                    pass

        return start_row + block_size + 1

    # Clear everything below row 7 to rebuild
    sheet.delete_rows(2, 500)
    
    # 2. Build Banners and Blocks
    row = 2
    # Banner Mayo
    sheet.cell(row=row, column=6).value = "📅 EXÁMENES FINAL - TURNO MAYO 2026"
    sheet.merge_cells(f"F{row}:Q{row}")
    # Styling for banner
    from openpyxl.styles import Font, PatternFill, Alignment
    banner_font = Font(size=14, bold=True, color="FFFFFF")
    banner_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    sheet.cell(row=row, column=6).font = banner_font
    sheet.cell(row=row, column=6).fill = banner_fill
    sheet.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    
    row += 2
    
    # Re-import template data validation needs to be handled carefully
    # Actually, we'll just use the list formulas directly
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_date = DataValidation(type="list", formula1="=LISTA_FECHAS", allow_blank=True)
    dv_time = DataValidation(type="list", formula1="=LISTA_HORARIOS", allow_blank=True)
    sheet.add_data_validation(dv_date)
    sheet.add_data_validation(dv_time)

    # ADD MAYO SUBJECTS
    # We'll use a local 'materia block' generator
    # But wait, copying styles is hard without a reference. 
    # Let's write the data first and then style one and copy it.
    
    # FOR SIMPLICITY: I will just generate the blocks manually in a loop with styles
    # based on the user's premium look.
    
    def write_styled_block(subject, info, call, r):
        # Basic Premium Styles
        font_name = Font(size=12, bold=True)
        font_info = Font(size=10, italic=True)
        fill_header = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        fill_call = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        fill_data = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Row 1: Name
        sheet.cell(row=r, column=6).value = subject
        sheet.merge_cells(start_row=r, start_column=6, end_row=r, end_column=17)
        sheet.cell(row=r, column=6).font = font_name
        sheet.cell(row=r, column=6).fill = fill_header
        
        # Row 2: Info
        sheet.cell(row=r+1, column=6).value = info
        sheet.merge_cells(start_row=r+1, start_column=6, end_row=r+1, end_column=17)
        sheet.cell(row=r+1, column=6).font = font_info
        
        # Row 3: Call + Countdown
        sheet.cell(row=r+2, column=6).value = call
        sheet.merge_cells(start_row=r+2, start_column=6, end_row=r+2, end_column=15)
        sheet.cell(row=r+2, column=6).fill = fill_call
        
        countdown_cell = sheet.cell(row=r+2, column=16)
        sheet.merge_cells(start_row=r+2, start_column=16, end_row=r+2, end_column=17)
        date_ref = f"O{r+3}"
        countdown_cell.value = f'=IF({date_ref}-TODAY()=0, "Hoy", IF({date_ref}-TODAY()=1, "Mañana", IF({date_ref}-TODAY()<0, "Finalizado", "En " & ({date_ref}-TODAY()) & " días")))'
        countdown_cell.alignment = Alignment(horizontal="center")
        countdown_cell.font = Font(bold=True)
        
        # Row 4: Fecha
        sheet.cell(row=r+3, column=12).value = "Fecha:"
        date_cell = sheet.cell(row=r+3, column=15)
        sheet.merge_cells(start_row=r+3, start_column=15, end_row=r+3, end_column=17)
        date_cell.fill = fill_data
        date_cell.number_format = 'DD-MM-YYYY'
        date_cell.alignment = Alignment(horizontal="center")
        dv_date.add(date_cell)
        
        # Row 5: Horario
        sheet.cell(row=r+4, column=12).value = "Horario:"
        time_cell = sheet.cell(row=r+4, column=15)
        sheet.merge_cells(start_row=r+4, start_column=15, end_row=r+4, end_column=17)
        time_cell.fill = fill_data
        time_cell.number_format = 'HH:mm'
        time_cell.alignment = Alignment(horizontal="center")
        dv_time.add(time_cell)
        
        return r + 7

    # Build Mayo
    for s, info in mesas_mayo:
        row = write_styled_block(s, info, "Llamado de Mayo", row)
        
    # Banner Coloquios
    row += 2
    sheet.cell(row=row, column=6).value = "🎓 COLOQUIOS / TALLERES - HISTORIA 2026"
    sheet.merge_cells(f"F{row}:Q{row}")
    coloquio_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    sheet.cell(row=row, column=6).font = banner_font
    sheet.cell(row=row, column=6).fill = coloquio_fill
    sheet.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    row += 2
    
    # Build Coloquios
    for s, info in coloquios:
        row = write_styled_block(s, info, "Coloquio", row)

    output_path = r'C:\Users\alanm\Desktop\wqwe\Tablero_Historia_Mayo_2026.xlsx'
    wb.save(output_path)
    print(f"Full Dashboard saved at: {output_path}")

if __name__ == "__main__":
    create_history_dashboard()
