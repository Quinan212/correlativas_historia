import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, time, timedelta

def create_high_compatibility_sheets_excel():
    file_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_PARA_SHEETS.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 1. Put data in the SAME sheet but far to the right (Column Z and AA)
    # Clear anything that might be there
    for r in range(1, 1000):
        sheet.cell(row=r, column=26).value = None # Z
        sheet.cell(row=r, column=27).value = None # AA
        
    # Dates in Column Z
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2027, 12, 31)
    current_date = start_date
    row_idx = 1
    while current_date <= end_date:
        sheet.cell(row=row_idx, column=26).value = current_date
        sheet.cell(row=row_idx, column=26).number_format = 'DD-MM-YYYY'
        current_date += timedelta(days=1)
        row_idx += 1
    max_date_row = row_idx - 1
        
    # Times in Column AA
    current_time = datetime.combine(datetime.today(), time(7, 0))
    end_time = datetime.combine(datetime.today(), time(22, 0))
    row_idx = 1
    while current_time <= end_time:
        sheet.cell(row=row_idx, column=27).value = current_time.time()
        sheet.cell(row=row_idx, column=27).number_format = 'HH:mm'
        current_time += timedelta(minutes=30)
        row_idx += 1
    max_time_row = row_idx - 1

    # 2. Use Direct Range Formulas (Google Sheets loves this)
    # Range for Dates: $Z$1:$Z$730
    # Range for Times: $AA$1:$AA$31
    
    dv_date = DataValidation(type="list", formula1=f"=$Z$1:$Z${max_date_row}", allow_blank=True)
    dv_time = DataValidation(type="list", formula1=f"=$AA$1:$AA${max_time_row}", allow_blank=True)
    
    sheet.add_data_validation(dv_date)
    sheet.add_data_validation(dv_time)
    
    # 3. Apply to all cells in the merge to ensure Sheets catches it
    # Merged cells were O6:P6, O10:P10, O16:P16 for dates
    # And O7:P7, O11:P11, O17:P17 for times (estimated)
    
    for r in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row=r, column=13).value).strip() if sheet.cell(row=r, column=13).value else ""
        if label == "Fecha:":
            dv_date.add(sheet.cell(row=r, column=15)) # O
            dv_date.add(sheet.cell(row=r, column=16)) # P (part of merge)
        elif label == "Horario:":
            dv_time.add(sheet.cell(row=r, column=15)) # O
            dv_time.add(sheet.cell(row=r, column=16)) # P
            
    # Optional: Hide columns Z and AA so they aren't visible
    sheet.column_dimensions['Z'].hidden = True
    sheet.column_dimensions['AA'].hidden = True
    
    output_path = r'C:\Users\alanm\Desktop\wqwe\eqwewq_COMPATIBLE_SHEETS.xlsx'
    wb.save(output_path)
    print(f"Compatible Excel for Sheets saved at: {output_path}")

if __name__ == "__main__":
    create_high_compatibility_sheets_excel()
