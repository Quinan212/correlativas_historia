"""
Analiza en profundidad los estilos actuales (fills, fonts, bordes) de las hojas
'MESAS', 'C.COLOQUIOS', 'Historia', 'Geografia', 'Ciencias Politicas' para
entender cómo ya están organizadas las "cards" y cómo rediseñarlas a un look
tipo dashboard estático (Google/Notion style).
"""
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import Counter

SRC = r"C:\Users\alanm\Desktop\Copia de mesas nuevo 800 VECES MEJOR.xlsx"
wb = load_workbook(SRC, data_only=False)

def fill_hex(cell):
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    fg = f.fgColor
    if fg is None:
        return None
    if fg.type == 'rgb' and fg.rgb:
        return fg.rgb
    if fg.type == 'theme':
        return f"theme:{fg.theme}"
    return str(fg.value)

def font_key(cell):
    fo = cell.font
    return (fo.name, fo.size, fo.bold, fo.color.rgb if fo.color and fo.color.type == 'rgb' else None)

for sh in ['MESAS', 'C.COLOQUIOS', 'Historia']:
    ws = wb[sh]
    print(f"\n===== {sh} =====")
    fills = Counter()
    fonts = Counter()
    rows_with_data = []
    for r in range(1, ws.max_row + 1):
        has_any = False
        row_info = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            fh = fill_hex(cell)
            fk = font_key(cell)
            fills[fh] += 1
            fonts[fk] += 1
            if cell.value is not None or fh not in (None, '00000000'):
                has_any = True
                if fh and fh != '00000000':
                    row_info.append((get_column_letter(c), fh, cell.value))
        if has_any and row_info:
            rows_with_data.append((r, row_info))

    print(f"Top fills:")
    for k, v in fills.most_common(12):
        print(f"  {k}: {v}")
    print(f"Top fonts:")
    for k, v in fonts.most_common(8):
        print(f"  {k}: {v}")
    print(f"Filas con fill activo (primeras 30):")
    for r, info in rows_with_data[:30]:
        preview = " ; ".join(f"{col}={fh}:{str(val)[:15] if val else ''}" for col, fh, val in info[:6])
        print(f"  R{r:03d}: {preview}")
