"""
Diagnóstico de precisión para encontrar el nombre exacto de la materia
"""
import openpyxl, os
CARPETA = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta'
ARCHIVO = os.path.join(CARPETA, 'Mesas Extraordinarias de Mayo Ciclo 2026 (22).xlsx')
wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
ws = wb['Historia']

print("Buscando patrones en Historia:")
for r in range(1, 150):
    c = ws.cell(r, 3).value # Col C
    d = ws.cell(r, 4).value # Col D
    j = ws.cell(r, 10).value # Col J
    k = ws.cell(r, 11).value # Col K
    
    # Si encontramos una fecha
    if "FECHA:" in str(j or "").upper() or "FECHA:" in str(k or "").upper():
        print(f"\n[R{r}] ENCONTRADA FECHA EN J={j} K={k}")
        print(f"      - Fila R: C='{c}', D='{d}'")
        print(f"      - Fila R-1: C='{ws.cell(r-1, 3).value}', D='{ws.cell(r-1, 4).value}'")
        print(f"      - Fila R-2: C='{ws.cell(r-2, 3).value}', D='{ws.cell(r-2, 4).value}'")
        print(f"      - Fila R-3: C='{ws.cell(r-3, 3).value}', D='{ws.cell(r-3, 4).value}'")
