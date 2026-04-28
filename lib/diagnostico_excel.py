"""
Comparar la hoja central 'Coloquios' con la sección de coloquios en 'Historia'
"""
import openpyxl, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ARCHIVO = os.path.join(r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta', 
                       'Mesas Extraordinarias de Mayo Ciclo 2026 (7).xlsx')

wb = openpyxl.load_workbook(ARCHIVO, data_only=True)

print("1. RESUMEN HOJA CENTRAL 'COLOQUIOS' (Fila 8-15):")
ws_c = wb['Coloquios']
for r in range(8, 15):
    line = [str(ws_c.cell(r, c).value or "").strip() for c in range(1, 15)]
    print(f"   R{r}: {line}")

print("\n2. RESUMEN HOJA 'HISTORIA' - SECCIÓN COLOQUIOS (Fila 125-140):")
ws_h = wb['Historia']
for r in range(125, 140):
    line = [str(ws_h.cell(r, c).value or "").strip() for c in range(1, 15)]
    print(f"   R{r}: {line}")
