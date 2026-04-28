import openpyxl, os

carpeta = r'C:\Users\alanm\Desktop\wqwe\Nueva carpeta'
archivos = [f for f in os.listdir(carpeta) if f.endswith('.xlsx')]

print("Buscando materias reales de Geografía en los archivos...")
for archivo in archivos:
    ruta = os.path.join(carpeta, archivo)
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for nombre_hoja in wb.sheetnames:
            if "GEOGRAFIA" in nombre_hoja.upper():
                ws = wb[nombre_hoja]
                # Escaneamos un poco de la hoja
                for row in ws.iter_rows(max_row=50, max_col=10):
                    for cell in row:
                        val = str(cell.value or "").upper()
                        if "GEOGRAFIA FISICA" in val or "CARTOGRAFIA" in val or "ESPACIOS MUNDIALES" in val:
                            print(f"\n¡ENCONTRADO! Archivo: {archivo} | Hoja: {nombre_hoja}")
                            print(f"Materia encontrada: {cell.value}")
                            wb.close()
                            exit()
        wb.close()
    except:
        continue
print("\nNo se encontraron las materias en los archivos principales.")
